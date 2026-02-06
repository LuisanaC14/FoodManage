from django.shortcuts import render, redirect
from django.db import models
from django.contrib.auth.models import User
from django.urls import path, reverse
from django.contrib import messages
from django.contrib import admin
from django.utils import timezone
from django.utils.html import format_html
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
from datetime import datetime, timedelta, time, date
from .models import ReservaPlato, Reserva, Asistencia, Venta, Mesa, Pedido, DetallePedido, Producto, Caja, SesionCaja, Gasto
from django import forms 
from .views import convertir_reserva_a_pedido
import csv  
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils.safestring import mark_safe
from urllib.parse import quote

def boton_editar(obj):
    url = reverse(f'admin:{obj._meta.app_label}_{obj._meta.model_name}_change', args=[obj.pk])
    return format_html(
        '<a class="button" href="{}" style="background:#28a745; color:white; padding:5px 12px; border-radius:5px; text-decoration:none; font-weight:bold; font-size:11px;">'
        '<i class="fas fa-edit me-1"></i>EDITAR</a>', 
        url
    )
boton_editar.short_description = "ACCIONES"
# =========================================================
# REGISTRO DE PRODUCTO (Base)
# =========================================================
@admin.register(Producto)
class ProductoAdmin(admin.ModelAdmin):
    change_form_template = 'admin/boton_volver_form.html'
    search_fields = ['nombre'] 
    list_display = ('nombre', 'precio', 'categoria') 

# =========================================================
# CONFIGURACIÓN DE LOS DETALLES (INLINE)
# =========================================================
class DetallePedidoInline(admin.TabularInline):
    model = DetallePedido
    extra = 0
    #autocomplete_fields = ['producto']
    fields = ('nota', 'producto', 'cantidad', 'precio_unitario', 'calcular_subtotal')
    readonly_fields = ('precio_unitario', 'calcular_subtotal')

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        field = super().formfield_for_foreignkey(db_field, request, **kwargs)
        
        if db_field.name == 'producto':
            from django.contrib.admin.widgets import RelatedFieldWidgetWrapper

            field.widget = RelatedFieldWidgetWrapper(
                field.widget,
                db_field.remote_field,
                self.admin_site,
                can_add_related=False,  
                can_change_related=False,
                can_delete_related=False, 
                can_view_related=False    
            )
            
        return field
    
    # 1. No permitir AGREGAR platos si ya pagó
    def has_add_permission(self, request, obj=None):
        if obj and obj.estado == 'Pagado':
            return False
        return True

    # 2. No permitir BORRAR platos si ya pagó
    def has_delete_permission(self, request, obj=None):
        if obj and obj.estado == 'Pagado':
            return False
        return True

    # 3. No permitir EDITAR cantidad/producto si ya pagó
    def has_change_permission(self, request, obj=None):
        if obj and obj.estado == 'Pagado':
            return False
        return True

    def calcular_subtotal(self, obj):
        precio = obj.precio_unitario if obj.precio_unitario is not None else 0
        cantidad = obj.cantidad if obj.cantidad is not None else 0
        return cantidad * precio
    calcular_subtotal.short_description = "Subtotal"

# 1. REGISTRO SIMPLE DE GASTOS (Para poder agregar gastos rápidos)
@admin.register(Gasto)
class GastoAdmin(admin.ModelAdmin):
    change_form_template = 'admin/boton_volver_form.html'
    # Agregamos 'boton_editar' al final de la lista
    list_display = ('concepto', 'monto_visual', 'categoria', 'fecha', 'usuario', 'boton_editar')
    list_filter = ('categoria', 'fecha')
    search_fields = ('concepto',)
    
    def has_delete_permission(self, request, obj=None):
        # Solo permite borrar si eres el Superusuario (Jefe)
        return request.user.is_superuser
    
    def monto_visual(self, obj):
        return format_html('<b style="color: #ff4d4d;">-${}</b>', obj.monto)
    monto_visual.short_description = "MONTO"

    # --- NUEVO: FUNCIÓN PARA EL BOTÓN EDITAR ---
    def boton_editar(self, obj):
        # Genera la URL para editar este gasto específico
        url = reverse('admin:gestion_gasto_change', args=[obj.id])
        return format_html(
            '<a class="btn-admin-edit" href="{}" style="font-size: 0.7rem; padding: 5px 10px;">'
            '<i class="fas fa-edit"></i> EDITAR'
            '</a>', 
            url
        )
    boton_editar.short_description = "ACCIÓN"
    boton_editar.allow_tags = True

# --- REEMPLAZA TODA LA CLASE VentaAdmin CON ESTO ---

@admin.register(Venta)
class VentaAdmin(admin.ModelAdmin):
    change_list_template = 'admin/gestion/venta/dashboard_ventas.html'

    def has_add_permission(self, request):
        return False

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('abrir-caja/', self.admin_site.admin_view(self.abrir_caja), name='venta_abrir_caja'),
            path('cerrar-caja/', self.admin_site.admin_view(self.cerrar_caja), name='venta_cerrar_caja'),
            path('exportar-excel/', self.admin_site.admin_view(self.exportar_excel), name='venta_excel'),
            path('nuevo-gasto/', self.admin_site.admin_view(self.nuevo_gasto_rapido), name='venta_nuevo_gasto'),
            path('imprimir-reporte/', self.admin_site.admin_view(self.vista_impresion), name='venta_imprimir'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        # --- CAMBIO AQUÍ: SIEMPRE ES HOY (Se borró el buscador) ---
        hoy = timezone.localdate()
        
        # Reservas futuras
        proximas_reservas = Reserva.objects.filter(
            fecha__gte=hoy,
            estado__in=['Pendiente', 'Confirmada']
        ).order_by('fecha', 'hora')[:5] 

        # ESTADO DE CAJA
        # Buscamos si hay una sesión abierta actualmente
        sesion_abierta = SesionCaja.objects.filter(estado='Abierta').last()
        caja_activa = True if sesion_abierta else False
        monto_inicial = sesion_abierta.monto_inicial if sesion_abierta else 0
            
        # FILTROS (Siempre Hoy)
        ventas = Venta.objects.filter(fecha_venta__date=hoy).order_by('-fecha_venta')
        gastos = Gasto.objects.filter(fecha__date=hoy).order_by('-fecha')
        
        total_ingresos = ventas.aggregate(Sum('total'))['total__sum'] or 0
        total_gastos = gastos.aggregate(Sum('monto'))['monto__sum'] or 0

        # DESGLOSE DE DINERO
        total_efectivo = ventas.filter(metodo_pago='Efectivo').aggregate(Sum('total'))['total__sum'] or 0
        total_transferencia = ventas.filter(metodo_pago__in=['Transferencia', 'Banco']).aggregate(Sum('total'))['total__sum'] or 0
        
        # Dinero en Gaveta = Fondo Inicial + Ventas Efectivo - Gastos
        dinero_en_caja = (monto_inicial + total_efectivo) - total_gastos

        # Top Productos
        top_productos = Venta.objects.filter(fecha_venta__date=hoy)\
            .values('producto__nombre')\
            .annotate(total_vendido=Sum('cantidad'))\
            .order_by('-total_vendido')[:5]

        # Gráfica
        datos_grafica = {}
        for v in ventas:
            hora = v.fecha_venta.strftime("%H:00")
            datos_grafica[hora] = datos_grafica.get(hora, 0) + float(v.total)
        chart_labels = list(datos_grafica.keys())
        chart_data = list(datos_grafica.values())

        if not chart_labels:
            chart_labels = [f"{h:02d}:00" for h in range(8, 24)]
            chart_data = [0] * len(chart_labels)

        utilidad_neta = total_ingresos - total_gastos

        context = {
            'title': f'REPORTE DEL DÍA: {hoy.strftime("%d/%m/%Y")}',
            'fecha_seleccionada': hoy.strftime("%Y-%m-%d"),
            'caja_abierta': caja_activa, 
            'monto_inicial': monto_inicial,
            'total_ingresos': total_ingresos,
            'total_gastos': total_gastos,
            'utilidad_neta': utilidad_neta,
            'total_efectivo': total_efectivo,
            'total_transferencia': total_transferencia,
            'dinero_en_caja': dinero_en_caja,
            'todas_ventas': ventas, 
            'todos_gastos': gastos,
            'top_productos': top_productos,
            'proximas_reservas': proximas_reservas,
            'chart_labels': chart_labels,
            'chart_data': chart_data,
        }
        context.update(extra_context or {})
        return super().changelist_view(request, extra_context=context)

    # --- ACCIONES ---
    def abrir_caja(self, request):
        if request.method == 'POST':
            monto = request.POST.get('monto', 0)
            SesionCaja.objects.create(usuario=request.user, monto_inicial=monto, estado='Abierta', fecha_apertura=timezone.now())
            messages.success(request, f"✅ Caja abierta con ${monto}")
        return redirect('admin:gestion_venta_changelist')

    def cerrar_caja(self, request):
        sesion = SesionCaja.objects.filter(estado='Abierta').last()
        if sesion:
            sesion.estado = 'Cerrada'
            sesion.fecha_cierre = timezone.now()
            sesion.save()
            messages.warning(request, "🔒 Turno cerrado.")
        return redirect('admin:gestion_venta_changelist')

    def nuevo_gasto_rapido(self, request):
        if request.method == 'POST':
            concepto = request.POST.get('concepto')
            monto = request.POST.get('monto')
            if concepto and monto:
                Gasto.objects.create(usuario=request.user, concepto=concepto, monto=monto)
                messages.success(request, f"💸 Gasto registrado: -${monto}")
        return redirect('admin:gestion_venta_changelist')

    def exportar_excel(self, request):
        from django.db.models import Sum, Count
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Caja.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cierre de Caja"
        
        # --- ESTILOS ---
        font_titulo = Font(bold=True, size=14, color="FFFFFF")
        fill_titulo = PatternFill("solid", fgColor="000000")
        font_bold = Font(bold=True)
        fill_verde = PatternFill("solid", fgColor="CCFF00") 
        fill_rojo = PatternFill("solid", fgColor="FF4D4D")
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        hoy = timezone.localdate()
        
        # 1. ENCABEZADO PRINCIPAL
        ws.merge_cells('A1:D1')
        ws['A1'] = f"REPORTE DE CAJA - {hoy}"
        ws['A1'].font = font_titulo
        ws['A1'].fill = fill_titulo
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # --- CÁLCULOS FINANCIEROS ---
        ventas_hoy = Venta.objects.filter(fecha_venta__date=hoy)
        total_v = ventas_hoy.aggregate(Sum('total'))['total__sum'] or 0
        total_efectivo = ventas_hoy.filter(metodo_pago='Efectivo').aggregate(Sum('total'))['total__sum'] or 0
        total_transf = ventas_hoy.filter(metodo_pago__in=['Transferencia', 'Banco']).aggregate(Sum('total'))['total__sum'] or 0
        
        gastos = Gasto.objects.filter(fecha__date=hoy)
        total_g = gastos.aggregate(Sum('monto'))['monto__sum'] or 0
        
        # 2. RESUMEN DE DINERO (Esto lo dejamos igual, es útil)
        ws['A3'] = "BALANCE DEL DÍA"
        ws['A3'].font = font_bold
        
        ws.append(["INGRESOS TOTALES", float(total_v)])
        ws.append(["   ↳ Efectivo", float(total_efectivo)])
        ws.append(["   ↳ Transferencia", float(total_transf)])
        ws.append(["GASTOS OPERATIVOS", float(total_g)])
        ws.append(["UTILIDAD NETA", float(total_v - total_g)])
        
        # Pintamos la utilidad
        ws['B7'].font = font_bold
        
        # 3. DETALLE DE GASTOS (Si hubo salidas de dinero)
        ws.append([]) # Espacio
        if gastos.exists():
            ws.append(["--- DESGLOSE DE GASTOS ---"])
            ws['A9'].font = font_bold
            ws.append(["Hora", "Concepto", "Monto"])
            for g in gastos:
                row = [g.fecha.strftime("%H:%M"), g.concepto, float(g.monto * -1)]
                ws.append(row)
                ws[f'C{ws.max_row}'].font = Font(color="FF0000")

        # 4. CONSOLIDADO DE PRODUCTOS (¡AQUÍ ESTÁ EL CAMBIO!)
        ws.append([]) # Espacio
        ws.append(["--- PRODUCTOS VENDIDOS (CONSOLIDADO) ---"])
        cell_titulo = ws[f'A{ws.max_row}']
        cell_titulo.font = font_bold
        
        # Encabezados de la tabla
        ws.append(["PRODUCTO", "CANTIDAD VENDIDA", "TOTAL GENERADO"])
        header_row = ws[ws.max_row]
        for cell in header_row:
            cell.font = font_bold
            cell.fill = fill_verde
            cell.border = border_thin

        # Lógica de Agrupación: Usamos values() y annotate() de Django
        # Esto hace la magia de convertir 50 filas en 1 sola
        productos_agrupados = ventas_hoy.values('producto__nombre').annotate(
            cantidad_total=Sum('cantidad'),
            dinero_total=Sum('total')
        ).order_by('-cantidad_total') # Ordenamos por el más vendido

        for p in productos_agrupados:
            ws.append([
                p['producto__nombre'], 
                p['cantidad_total'], 
                float(p['dinero_total'])
            ])
            # Bordes simples
            for cell in ws[ws.max_row]: cell.border = border_thin

        # Ajuste de ancho de columnas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        wb.save(response)
        return response

    def vista_impresion(self, request):
        hoy = timezone.localdate()

        ventas = Venta.objects.filter(fecha_venta__date=hoy).order_by('fecha_venta')
        gastos = Gasto.objects.filter(fecha__date=hoy).order_by('fecha')
        
        t_ingresos = ventas.aggregate(Sum('total'))['total__sum'] or 0
        t_gastos = gastos.aggregate(Sum('monto'))['monto__sum'] or 0
        
        # Busca sesión abierta o la última del día
        sesion = SesionCaja.objects.filter(fecha_apertura__date=hoy).last()
        inicio = sesion.monto_inicial if sesion else 0
        
        context = {
            'fecha': hoy, 
            'usuario': request.user.username, 
            'monto_inicial': inicio,
            'total_ingresos': t_ingresos, 
            'total_gastos': t_gastos, 
            'utilidad_neta': t_ingresos - t_gastos,
            'ventas': ventas, 
            'gastos': gastos,
        }
        return render(request, 'admin/gestion/venta/reporte_impresion.html', context)
    
@admin.register(Caja)
class CajaAdmin(admin.ModelAdmin):
    # Esta usa la plantilla OSCURA original (asegúrate de que ese archivo tenga el HTML del POS)
    # No definimos change_list_template aquí porque usamos una vista personalizada
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('', self.admin_site.admin_view(self.vista_cajero), name='caja_dashboard'),
            path('cobrar/<int:pedido_id>/', self.admin_site.admin_view(self.procesar_cobro), name='caja_cobrar'),
        ]
        return custom_urls + urls

    def vista_cajero(self, request):
        context = {
            **self.admin_site.each_context(request),
            'title': 'Caja y Cobros - Máncora Marisquería',
            'pedidos': Pedido.objects.filter(estado__in=['Pendiente', 'En preparación', 'Listo']).order_by('-fecha_pedido'),
            'mesas': Mesa.objects.all(),
        }

        hoy = timezone.now().date()
        
        # FILTRO LIMPIO: Solo muestra lo que falta por cobrar
        pedidos_activos = Pedido.objects.filter(
            estado__in=['Pendiente', 'En preparación', 'Listo']
        ).order_by('-fecha_pedido')

        context = {
            **self.admin_site.each_context(request),
            'title': 'Caja y Cobros - Máncora Marisquería',
            'pedidos': pedidos_activos, # Usamos la lista filtrada
            'mesas': Mesa.objects.all(),
        }
        return render(request, 'admin/gestion/caja/dashboard.html', context)
        # IMPORTANTE: Este HTML debe ser el de tu diseño "PEDIDOS POR COBRAR" (El oscuro)
        return render(request, 'admin/gestion/caja/dashboard.html', context)

    # En gestion/admin.py dentro de CajaAdmin

    def procesar_cobro(self, request, pedido_id):
        if request.method == 'POST':
            try:
                pedido = Pedido.objects.get(id=pedido_id)
                
                # 1. CAPTURAR MÉTODO DE PAGO (Lo que arreglamos antes)
                metodo = request.POST.get('metodo_pago')
                if not metodo: 
                    metodo = 'Efectivo'

                # 2. CAPTURAR DATOS DEL CLIENTE (¡AQUÍ ESTÁ EL ARREGLO!)
                # Leemos lo que escribiste en los inputs del Dashboard
                nuevo_nombre = request.POST.get('factura-nombre')
                nuevo_ruc = request.POST.get('factura-id')
                nuevo_telf = request.POST.get('factura-telefono')
                nuevo_email = request.POST.get('factura-email')
                nueva_dir = request.POST.get('factura-direccion')

                # Si el cajero escribió algo, actualizamos el pedido
                if nuevo_nombre:
                    pedido.cliente_nombre = nuevo_nombre
                if nuevo_ruc:
                    pedido.cliente_cedula = nuevo_ruc
                if nuevo_telf:
                    pedido.cliente_telefono = nuevo_telf
                if nuevo_email:
                    pedido.cliente_email = nuevo_email
                if nueva_dir:
                    pedido.cliente_direccion = nueva_dir
                
                # 3. CREAR LAS VENTAS PARA EL REPORTE
                for d in pedido.detalles.all():
                    Venta.objects.create(
                        producto=d.producto,
                        cantidad=d.cantidad,
                        total=d.cantidad * d.precio_unitario,
                        metodo_pago=metodo,
                        fecha_venta=timezone.now()
                    )
                
                # 4. GUARDAR CAMBIOS Y FINALIZAR
                pedido.estado = 'Pagado'
                pedido.save() # Aquí se guardan los datos del cliente y el estado
                
                messages.success(request, f"¡Cobro exitoso a {pedido.cliente_nombre}!")
                
            except Exception as e:
                messages.error(request, f"Error: {e}")
                
            return redirect('admin:caja_dashboard')
    

# =========================================================
# ADMIN DE PEDIDOS
# =========================================================
@admin.register(Pedido)
class PedidoAdmin(admin.ModelAdmin):
    change_form_template = 'admin/boton_volver_form.html'
    change_form_template = 'admin/gestion/pedido/change_form.html'
    inlines = [DetallePedidoInline]
    
    # UNIFICAMOS list_display para incluir al cliente y los botones
    list_display = ('numero_visual', 'mesa_visual', 'cliente_nombre', 'cliente_cedula', 'cliente_telefono', 
        'cliente_email', 'cliente_direccion' ,'solo_cantidades','resumen_productos', 'resumen_notas', 'fecha_bonita', 'estado_visual', 
        'total_visual','metodo_pago','ver_comprobante', 'boton_imprimir', 'boton_email', 'boton_editar',
    )
    
    fields = (
        'mesa', 'mesero', 'estado', 'cliente_nombre', 'cliente_cedula','cliente_telefono','cliente_email', 'cliente_direccion', 'observaciones'
    )
    
    list_filter = ('estado', 'fecha_pedido','metodo_pago',) 
    search_fields = ('mesero__username', 'mesa__numero', 'cliente_nombre', 'cliente_cedula') 

    def has_delete_permission(self, request, obj=None):
        # Esto quita el botón de eliminar de la lista y del formulario
        return False

    def ver_comprobante(self, obj):
            if obj.comprobante_pago:
                return format_html(
                    '<a href="{}" target="_blank" style="background:#ccff00; color:black; padding:5px 10px; border-radius:50px; font-weight:bold; text-decoration:none;">'
                    '<i class="fas fa-image"></i> Ver Foto</a>',
                    obj.comprobante_pago.url
                )
            elif obj.metodo_pago == 'Transferencia':
                return mark_safe('<li class="page-item disabled"><span class="page-link">&hellip;</span></li>')
            else:
                return "-"
    
    ver_comprobante.short_description = "Comprobante"

    # --- 2. NUEVA FUNCIÓN VISUAL (ESTILO CÍRCULO) ---
    def numero_visual(self, obj):
        # Usamos colores según el estado para que sea más útil visualmente
        if obj.estado == 'Pendiente':
            color = "#ffae00" # Naranja
        elif obj.estado == 'Listo':
            color = "#ccff00" # Verde Neón
        elif obj.estado == 'Pagado':
            color = "#00eaff" # Azul Cyan
        else:
            color = "#666" # Gris

        return format_html(
            '<div style="width: 35px; height: 35px; background: #1a1a1a; border: 2px solid {}; '
            'border-radius: 50%; display: flex; align-items: center; justify-content: center; '
            'color: #fff; font-weight: 900; font-size: 0.9rem; box-shadow: 0 0 5px rgba(0,0,0,0.5); margin: 0 auto;">'
            '#{}</div>',
            color, obj.numero_diario
        )
    numero_visual.short_description = "ID"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            # Ruta para ver el monitor
            path('cocina/', self.admin_site.admin_view(self.vista_cocina), name='monitor_cocina'),
            # RUTA VITAL PARA EL BOTÓN LISTO (Asegúrate de que esta línea EXISTA)
            path('<int:pedido_id>/set_listo/', self.admin_site.admin_view(self.set_listo), name='pedido_set_listo'),
        ]
        return custom_urls + urls

    def boton_email(self, obj):
        # Usamos un campo de email. Si tu modelo Pedido NO tiene campo email, 
        # el botón aparecerá deshabilitado o gris.
        email = getattr(obj, 'cliente_email', '') # Cambia 'cliente_email' por el nombre real de tu campo
        
        if email:
            url = reverse('enviar_ticket_email', args=[obj.id])
            return format_html(
                '<a class="button" href="{}" style="background-color: #6c5ce7; color: white; padding: 4px 8px; border-radius: 4px; margin-left: 5px;" title="Enviar a {}">'
                '<i class="fas fa-envelope"></i></a>', 
                url, email
            )
        else:
            return mark_safe(
                '<span style="color: #666; cursor: not-allowed; padding: 4px 8px;" title="Sin correo registrado">'
                '<i class="fas fa-envelope-open"></i></span>'
            )
            
    boton_email.short_description = "ENVIAR"
    boton_email.allow_tags = True

    def boton_imprimir(self, obj):
        try:
            url = reverse('imprimir_ticket', args=[obj.id])
            # AQUÍ ESTÁ EL CAMBIO: width=1100, height=900 (Ventana Grande para Factura A4)
            return format_html(
                '<a class="btn btn-warning btn-xs" style="cursor: pointer;" '
                'onclick="window.open(\'{}\', \'TicketPrint\', \'width=1100,height=900,scrollbars=yes,resizable=yes\'); return false;" '
                'title="Imprimir Factura">'
                '<i class="fas fa-print"></i></a>', 
                url
            )
        except:
            return "No URL"
            
    boton_imprimir.short_description = "TICKET"
    boton_imprimir.allow_tags = True

    def set_listo(self, request, pedido_id):
        if request.method == 'POST':
            pedido = Pedido.objects.get(id=pedido_id)
            pedido.estado = 'Listo' # Aquí es donde se guarda el cambio real
            pedido.save()
            return JsonResponse({'status': 'ok'})

    def vista_cocina(self, request):
        pedidos = Pedido.objects.filter(estado__in=['Pendiente', 'En preparación']).order_by('fecha_pedido')
        # Traemos los pedidos pendientes con sus platos ya cargados
        pedidos = Pedido.objects.filter(
            estado__in=['Pendiente', 'En preparación']
        ).prefetch_related('detalles__producto').order_by('fecha_pedido')
        
        context = {
            **self.admin_site.each_context(request),
            'pedidos': pedidos,
            'title': 'MONITOR DE COCINA - MÁNCORA',
        }
        return render(request, 'admin/gestion/cocina/cocina.html', context)

    def get_readonly_fields(self, request, obj=None):
        campos_lectura = ['total'] 
        # Corregido de 'usuario' a 'mesero' según tu modelo previo
        if request.user.groups.filter(name='Meseros').exists():
            campos_lectura.extend(['mesero', 'mesa', 'estado'])
        return campos_lectura
    
    # --- FUNCIONES VISUALES ---

    def id_decorado(self, obj):
        from django.utils.html import format_html
        # Añadimos text-align: center y display: block al contenedor del número
        return format_html(
            '<div style="text-align: center; display: block; width: 100%; color: #666; font-size: 1.2rem;">'
            '<b>#{}</b>'
            '</div>', 
            obj.numero_diario
        )
    id_decorado.short_description = "TICKET"

    def mesa_visual(self, obj):
        from django.utils.html import format_html
        return format_html('<span class="text-mesa">Mesa {}</span>', obj.mesa.numero) if obj.mesa else "Sin Mesa"
    mesa_visual.short_description = "MESA"

    # 2. Función para mostrar solo los números en un apartado visual
    def solo_cantidades(self, obj):
        detalles = obj.detalles.all()
        agrupados = {}
        for item in detalles:
            nombre = item.producto.nombre
            agrupados[nombre] = agrupados.get(nombre, 0) + item.cantidad
        
        html = []
        for nombre, cantidad in agrupados.items():
            html.append(format_html(
                '<div style="margin-bottom: 8px; text-align: center;">'
                '<span style="background: #ccff00; color: #000; font-weight: 900; '
                'padding: 3px 10px; border-radius: 5px; display: inline-block; '
                'min-width: 28px; box-shadow: 0 0 5px rgba(204,255,0,0.5);">{}</span>'
                '</div>', cantidad
            ))
        return mark_safe("".join(html)) if html else "-"
    solo_cantidades.short_description = "CANT."

    # 3. Función para mostrar solo los nombres de los platos al lado
    def resumen_productos(self, obj):
        detalles = obj.detalles.all()
        # Sacamos solo los nombres únicos para que coincidan con las cantidades de al lado
        nombres = []
        vistos = set()
        for item in detalles:
            if item.producto.nombre not in vistos:
                nombres.append(item.producto.nombre)
                vistos.add(item.producto.nombre)
        
        html = []
        for nombre in nombres:
            html.append(format_html(
                '<div style="margin-bottom: 8px; color: #fff; font-weight: bold; '
                'height: 24px; display: flex; align-items: center; white-space: nowrap;">'
                '{}</div>', nombre
            ))
        return mark_safe("".join(html)) if html else "-"
    resumen_productos.short_description = "PRODUCTOS"

    def resumen_notas(self, obj):
        from django.utils.html import format_html
        from django.utils.safestring import mark_safe
        
        # 1. Primero creamos la lista y el set de memoria
        notas_html = []
        textos_vistos = set() 
        
        # 2. Mostramos notas de los platos y las guardamos en 'textos_vistos'
        for item in obj.detalles.all():
            if item.nota:
                # Usamos lower() para que "Sin Cebolla" y "sin cebolla" se consideren iguales
                nota_limpia = item.nota.strip().lower() 
                textos_vistos.add(nota_limpia)
                
                html = format_html(
                    '<div style="margin-bottom:4px;">'
                    '<span style="color:#888; font-size:0.85rem;">{}</span>: '
                    '<span style="color:#ccff00; font-weight:bold; text-transform:uppercase;">{}</span>'
                    '</div>',
                    item.producto.nombre,
                    item.nota
                )
                notas_html.append(html)
        
        # 3. Mostramos la nota general solo si no es un duplicado
        if obj.observaciones:
            obs_limpia = obj.observaciones.strip().lower()
            
            if obs_limpia not in textos_vistos:
                html_general = format_html(
                    '<div style="margin-top:5px; border-top:1px solid #333; padding-top:2px;">'
                    '<span style="color:#00bfff; font-size:0.8rem;">GENERAL:</span> '
                    '<span style="color:#fff;">{}</span>'
                    '</div>',
                    obj.observaciones
                )
                notas_html.append(html_general)

        if not notas_html:
            return "-"
            
        return mark_safe("".join(notas_html))

    resumen_notas.short_description = "NOTAS / DETALLES"

    def fecha_bonita(self, obj):
        return obj.fecha_pedido.strftime("%d/%m %H:%M") 
    fecha_bonita.short_description = "HORA"

    def estado_visual(self, obj):
        from django.utils.html import format_html
        if obj.estado == 'Pendiente' or obj.estado == 'En preparación':
            clase = 'status-pendiente'
            texto = obj.estado.upper()
        else:
            clase = 'status-completado'
            texto = obj.estado.upper()
        return format_html('<span class="badge-status {}">{}</span>', clase, texto)
    estado_visual.short_description = "ESTADO"

    def total_visual(self, obj):
        from django.utils.html import format_html
        return format_html('<span class="text-total">${}</span>', obj.total)
    total_visual.short_description = "TOTAL"

    def boton_editar(self, obj):
        from django.utils.html import format_html
        return format_html('<a class="btn-admin-edit" href="/admin/gestion/pedido/{}/change/"><i class="fas fa-edit"></i> EDITAR</a>', obj.id)
    boton_editar.short_description = "EDITAR"

    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        filtro = ~Q(categoria__icontains='bebida') | Q(categoria__icontains='bebida', stock__gt=0)
        # Aplicamos el filtro
        context['productos_visuales'] = Producto.objects.filter(filtro).order_by('categoria', 'nombre')
        return super().render_change_form(request, context, add, change, form_url, obj)
    
    def get_readonly_fields(self, request, obj=None):
        # Campos que SIEMPRE son de lectura (cálculos)
        campos_bloqueados = ['total']
        
        # Si el usuario es Mesero, bloqueamos mesa y estado también (tu lógica actual)
        if request.user.groups.filter(name='Meseros').exists():
            campos_bloqueados.extend(['mesero', 'mesa', 'estado'])

        if obj and obj.estado == 'Pagado':
            # Bloqueamos TODO lo operativo
            campos_bloqueados.extend(['mesa', 'mesero', 'estado', 'observaciones'])
            # NOTA: NO bloqueamos 'cliente_nombre', 'cliente_cedula', etc.
            # Así podrás seguir editando la factura sin alterar el pedido.
            
        return campos_bloqueados 
    
    list_per_page = 10  # Muestra solo 20 pedidos por página para mayor rapidez
    
    def get_queryset(self, request):
        """Filtra para que al entrar solo veas lo importante"""
        qs = super().get_queryset(request)
        
        # Si el usuario NO ha aplicado ningún filtro manual (está entrando recién)
        if not request.GET:
            # Solo muestra pedidos que NO estén Pagados ni Cancelados
            return qs.exclude(estado__in=['Pagado', 'Cancelado'])
        
        return qs


@admin.register(Asistencia)
class AsistenciaAdmin(admin.ModelAdmin):
    change_list_template = 'admin/gestion/asistencia/change_list_asistencia.html'
    change_form_template = 'admin/boton_volver_form.html'

    list_display = ('empleado_nombre', 'dia_semana', 'fecha_bonita', 'hora_exacta', 'estado_inteligente', 'nota_visual')
    list_filter = ('fecha', 'empleado')
    search_fields = ('empleado__username',)
    ordering = ('-fecha', 'hora_entrada') 
    fields = ('empleado', 'nota') 

    def has_add_permission(self, request):
        return request.user.is_superuser or request.user.groups.filter(name='Cajeros').exists()

    # --- 1. LÓGICA DEL DASHBOARD (TARJETAS ARRIBA) ---
    def changelist_view(self, request, extra_context=None):
        hoy = timezone.localdate()
        
        # Filtramos asistencias de HOY
        qs_hoy = Asistencia.objects.filter(fecha=hoy)
        
        total_presentes = qs_hoy.count()
        
        # Calculamos Atrasos: Cualquier hora MAYOR a las 08:00:00
        hora_limite = time(8, 0, 0)
        total_atrasos = qs_hoy.filter(hora_entrada__gt=hora_limite).count()
        
        total_puntuales = total_presentes - total_atrasos

        extra_context = extra_context or {}
        extra_context['dashboard_asistencia'] = {
            'presentes': total_presentes,
            'atrasos': total_atrasos,
            'puntuales': total_puntuales,
        }
        return super().changelist_view(request, extra_context=extra_context)

    # --- 2. ESTADO INTELIGENTE (VERDE vs ROJO) ---
    def estado_inteligente(self, obj):
        # HORA LÍMITE: 08:00 AM EXACTA
        hora_limite = time(8, 0, 0)
        
        if obj.hora_entrada > hora_limite:
            # SI ES TARDE (ROJO)
            return mark_safe(
                '<span style="background: rgba(255, 77, 77, 0.1); color: #ff4d4d; border: 1px solid #ff4d4d; '
                'padding: 3px 10px; border-radius: 4px; font-size: 0.8rem; font-weight: bold;">'
                '<i class="fas fa-exclamation-circle"></i> TARDE</span>'
            )
        else:
            # SI ES PUNTUAL (VERDE)
            return mark_safe(
                '<span style="background: rgba(40, 167, 69, 0.1); color: #28a745; border: 1px solid #28a745; '
                'padding: 3px 10px; border-radius: 4px; font-size: 0.8rem; font-weight: bold;">'
                '<i class="fas fa-check"></i> PUNTUAL</span>'
            )
    estado_inteligente.short_description = "ESTADO"

    def nota_visual(self, obj):
        if obj.nota:
            return mark_safe(f'<span style="color:#aaa; font-style:italic;">"{obj.nota}"</span>')
        return "-"
    nota_visual.short_description = "NOTA / EXCUSA"

    # --- TUS MÉTODOS VISUALES ORIGINALES ---
    def empleado_nombre(self, obj):
        return mark_safe(f'<span style="color: #00eaff; font-weight: 900; font-size: 1.1rem;">{obj.empleado.username.upper()}</span>')
    empleado_nombre.short_description = "EMPLEADO"

    def dia_semana(self, obj):
        dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        return mark_safe(f'<span style="color: #fff; opacity:0.7;">{dias[obj.fecha.weekday()]}</span>')
    dia_semana.short_description = "DÍA"

    def fecha_bonita(self, obj):
        return obj.fecha.strftime("%d/%m/%Y")
    fecha_bonita.short_description = "FECHA"

    def hora_exacta(self, obj):
        # Resaltamos la hora para ver los segundos
        return mark_safe(f'<span style="color: #ccff00; font-family: monospace; font-size: 1.2rem;">{obj.hora_entrada.strftime("%H:%M:%S")}</span>')
    hora_exacta.short_description = "HORA REGISTRO"

    # --- CONFIGURACIÓN DE FORMULARIO Y GUARDADO (INTACTOS) ---
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'empleado' in form.base_fields:
            form.base_fields['empleado'].queryset = User.objects.filter(is_active=True).order_by('username')
            form.base_fields['empleado'].label = "SELECCIONE AL EMPLEADO"
            form.base_fields['empleado'].widget.can_add_related = False
            form.base_fields['empleado'].widget.can_change_related = False
            form.base_fields['empleado'].widget.can_view_related = False
            form.base_fields['empleado'].widget.can_delete_related = False

        if 'nota' in form.base_fields:
            form.base_fields['nota'].widget.attrs.update({
                'placeholder': 'Escriba el motivo si llega tarde...',
                'style': 'width: 100%; border-color: #555;'
            })
            form.base_fields['nota'].help_text = "Opcional: Indique el motivo del retraso si registra después de las 08:00 AM."
        return form

    def save_model(self, request, obj, form, change):
        # 1. Validar doble asistencia
        if not change:
            ya_marco = Asistencia.objects.filter(
                empleado=obj.empleado,
                fecha=timezone.now().date()
            ).exists()
            if ya_marco:
                messages.error(request, f"⚠️ ERROR: {obj.empleado.username.upper()} ya registró su asistencia hoy.")
                return 

        # 2. Advertencia si es tarde (para que el que registra sepa)
        hora_actual = timezone.localtime(timezone.now()).time()
        if hora_actual > time(8, 0) and not obj.nota:
            messages.warning(request, "⚠️ Registro después de las 08:00 AM. Se recomienda indicar el motivo del retraso.")
        
        super().save_model(request, obj, form, change)

@admin.register(Mesa)
class MesaAdmin(admin.ModelAdmin):
    change_list_template = 'admin/gestion/mesa/change_list.html'
    change_form_template = 'admin/boton_volver_form.html'
    
    # Columnas limpias (Sin íconos extraños)
    list_display = ('numero_visual', 'piso_visual', 'capacidad_visual', 'forma', 'pos_x', 'pos_y', 'boton_editar')
    
    list_editable = ('pos_x', 'pos_y', 'forma') 
    list_filter = ('piso', 'capacidad')
    ordering = ('piso', 'numero')

    # --- 1. ESTILOS DE LOS INPUTS (CORREGIDO: ADIÓS ROJO) ---
    formfield_overrides = {
        models.IntegerField: {
            'widget': forms.TextInput(attrs={
                'style': (
                    'width: 60px; text-align: center; '
                    'background: #111; color: #00eaff; font-weight: bold; '
                    'border: 1px solid #444; border-radius: 4px; padding: 6px; outline: none;'
                )
            })
        },
        models.CharField: {
            'widget': forms.Select(attrs={
                'style': (
                    'width: 170px; background: #111; color: #fff; '
                    'border: 1px solid #444; border-radius: 4px; padding: 6px; outline: none;'
                )
            })
        }, 
    }

    # --- MÉTODOS VISUALES ---

    def numero_visual(self, obj):
        color_borde = "#ffcc00" if obj.piso == "Piso 1" else "#00eaff"
        return format_html(
            '<div style="width: 35px; height: 35px; background: #1a1a1a; border: 2px solid {}; '
            'border-radius: 50%; display: flex; align-items: center; justify-content: center; '
            'color: #fff; font-weight: 900; font-size: 0.9rem; box-shadow: 0 0 8px rgba(0,0,0,0.5);">'
            '{}</div>',
            color_borde, obj.numero
        )
    numero_visual.short_description = "N°"

    def piso_visual(self, obj):
        color = "#ffcc00" if obj.piso == "Piso 1" else "#00eaff"
        texto = "PLANTA BAJA" if obj.piso == "Piso 1" else "TERRAZA"
        icono = "fa-layer-group" if obj.piso == "Piso 1" else "fa-cloud-sun"
        
        return format_html(
            '<span style="color: {}; font-weight: bold; font-size: 0.8rem; letter-spacing: 0.5px;">'
            '<i class="fas {}" style="margin-right:5px; opacity:0.7;"></i>{}'
            '</span>',
            color, icono, texto
        )
    piso_visual.short_description = "UBICACIÓN"

    def capacidad_visual(self, obj):
        return format_html(
            '<div style="color: #aaa; font-size: 0.85rem;">'
            '<i class="fas fa-user-friends" style="color: #666; margin-right: 5px;"></i>'
            '<b>{}</b> pers.'
            '</div>', 
            obj.capacidad
        )
    capacidad_visual.short_description = "CAPACIDAD"

    # --- 2. BOTÓN EDITAR 
    def boton_editar(self, obj):
        url = reverse('admin:gestion_mesa_change', args=[obj.id])
        return format_html(
            '<a class="btn-admin-edit" href="{}" style="background: #222; border: 1px solid #444; color: #00eaff; padding: 6px 12px; border-radius: 6px; display: inline-block; transition: 0.3s;" title="Editar Mesa">'
            '<i class="fas fa-pen" style="font-size: 1rem;"></i>'
            '</a>', 
            url
        )
    boton_editar.short_description = "EDITAR"

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['title'] = 'Configuracion de Mesas' 
        return super().changelist_view(request, extra_context=extra_context)

class ReservaPlatoInline(admin.TabularInline):
    model = ReservaPlato
    extra = 1  # Muestra 1 fila vacía lista para llenar
    min_num = 0
    
    # Esto busca el plato escribiendo su nombre (requiere search_fields en ProductoAdmin)
    autocomplete_fields = ['producto'] 
    
    verbose_name = "Plato para adelantar"
    verbose_name_plural = "🍽️ PRE-ORDEN DE COMIDA (Opcional)"
    
    formfield_overrides = {
        models.CharField: {
            'widget': forms.TextInput(attrs={
                'placeholder': 'Ej: Sin cebolla, Salsa aparte...',
                'style': 'width: 300px;' # Hacemos el campo más ancho
            })
        },
    }

@admin.register(Reserva)
class ReservaAdmin(admin.ModelAdmin):
    change_list_template = 'admin/gestion/reserva/change_list_custom.html'
    change_form_template = 'admin/gestion/reserva/change_form.html'

    # Agregamos 
    list_display = ('alerta_visual', 'cliente_neon', 'boton_whatsapp', 'piso_badge', 'mesa_visual', 'fecha_hora', 'personas_visual', 'estado_color', 'boton_editar', 'boton_convertir')
    list_filter = ('mesa__piso', 'estado', 'fecha')
    search_fields = ('cliente', 'mesa__numero', 'telefono') # Agregamos búsqueda por teléfono
    ordering = ('fecha', 'hora')
    
    # Agregamos los nuevos campos al formulario de edición
    fields = ['cliente', 'telefono', 'numero_personas', 'fecha', 'hora', 'mesa', 'asistio', 'estado', 'notas']
    inlines = [ReservaPlatoInline]

    def has_delete_permission(self, request, obj=None):
        # Solo el Administrador Principal (Superuser) puede borrar reservas
        # Los empleados deben cambiar el estado a "Cancelada"
        return request.user.is_superuser
    # --- 1. LÓGICA DEL PANEL DE RESUMEN (VISTAZO RÁPIDO) ---
    def changelist_view(self, request, extra_context=None):
        hoy = timezone.localdate()
        
        # Filtramos solo las de HOY que no estén canceladas
        qs_hoy = Reserva.objects.filter(fecha=hoy).exclude(estado='Cancelada')
        
        # Calculamos métricas
        total_reservas = qs_hoy.count()
        # Sumamos el campo 'numero_personas'
        total_personas = qs_hoy.aggregate(Sum('numero_personas'))['numero_personas__sum'] or 0
        pendientes = qs_hoy.filter(estado='Pendiente').count()

        extra_context = extra_context or {}
        extra_context['dashboard_reservas'] = {
            'total': total_reservas,
            'personas': total_personas,
            'pendientes': pendientes,
            'hoy': hoy,
        }
        return super().changelist_view(request, extra_context=extra_context) 

    # --- 2. BOTÓN WHATSAPP AUTOMÁTICO ---
    def boton_whatsapp(self, obj):
        if not obj.telefono:
            return "-"
        
        # Limpieza del número
        numero = obj.telefono.replace(" ", "").replace("-", "")
        if numero.startswith("0"):
            numero = "593" + numero[1:]
        
        # Mensaje con estructura de saltos de línea
        mensaje_texto = (
            f"ESTIMADO(A): *{obj.cliente.upper()}*\n\n"
            f"Reciba un cordial saludo de *Máncora Marisquería*.\n"
            f"Por medio del presente, confirmamos los detalles de su reserva para el día de hoy:\n\n"
            f"--- DETALLES DE RESERVACIÓN ---\n"
            f"*HORA:* {obj.hora.strftime('%H:%M')}\n"
            f"*PERSONAS:* {obj.numero_personas}\n"
            f"*UBICACIÓN:* {obj.mesa.piso.upper()}\n"
            f"-------------------------------\n\n"
            f"Estamos preparando todo para brindarle una experiencia excepcional.\n"
            f"Si desea adelantar su pedido de platos fuertes o requiere asistencia adicional, por favor háganoslo saber por este medio.\n\n"
            f"¡Le esperamos!"
        )

        # La clave: Convertir el texto a formato de URL segura
        mensaje_url = quote(mensaje_texto)
        url = f"https://wa.me/{numero}?text={mensaje_url}"
        
        return format_html(
            '<a href="{}" target="_blank" style="color: #25D366; font-size: 1.2rem; text-decoration: none;" title="Enviar Confirmación">'
            '<i class="fab fa-whatsapp"></i>'
            '</a>',
            url
        )
    
    boton_whatsapp.short_description = "WP"

    # --- 3. VISUALIZACIÓN DE PERSONAS ---
    def personas_visual(self, obj):
        return format_html(
            '<div style="text-align:center;">'
            '<i class="fas fa-user-friends" style="color:#888; margin-right:5px;"></i>'
            '<b style="color:#fff; font-size:1rem;">{}</b>'
            '</div>', 
            obj.numero_personas
        )
    personas_visual.short_description = "#PERSONAS"

    # --- TUS FUNCIONES ANTERIORES (INTACTAS) ---
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [path('<int:reserva_id>/convertir/', self.admin_site.admin_view(convertir_reserva_a_pedido), name='reserva_convertir')]
        return custom_urls + urls

    def boton_convertir(self, obj):
        if not obj.asistio and obj.estado != 'Cancelada':
            url = reverse('admin:reserva_convertir', args=[obj.id])
            return format_html('<a class="btn-admin-go" href="{}" title="Cliente Llegó"><i class="fas fa-rocket"></i> ACTIVAR</a>', url)
        elif obj.asistio:
             return mark_safe('<span style="color:#666; font-size:0.8rem;">YA ASISTIÓ</span>')
        else: return "-"
    boton_convertir.short_description = "LLEGADA"
    boton_convertir.allow_tags = True

    def boton_editar(self, obj):
        url = reverse('admin:gestion_reserva_change', args=[obj.id])
        return format_html('<a class="btn-admin-edit" href="{}"><i class="fas fa-edit"></i></a>', url)
    boton_editar.short_description = "EDIT"
    boton_editar.allow_tags = True

    def alerta_visual(self, obj):
        hoy = timezone.now().date()
        dias_faltantes = (obj.fecha - hoy).days
        if obj.estado in ['Cancelada', 'Finalizada']:
            return mark_safe('<div style="color:#666; font-size:18px; text-align:center;"><i class="fas fa-check-square"></i></div>')
        if dias_faltantes < 0: return mark_safe('<div style="color:#555; font-size:18px; text-align:center;"><i class="fas fa-history"></i></div>')
        elif dias_faltantes == 0: return mark_safe('<div style="color:#ff0000; font-weight:900; animation: parpadeo 1s infinite; text-align:center; border: 1px solid #ff0000; padding: 2px; border-radius: 4px; font-size:0.7rem;">¡HOY!</div>')
        elif dias_faltantes == 1: return mark_safe('<div style="color:#ffae00; text-align:center; font-weight:bold; font-size:0.7rem;">MAÑANA</div>')
        else: return mark_safe(f'<div style="color:#00eaff; text-align:center; font-weight:bold; font-size:0.7rem;">En {dias_faltantes} días</div>')
    alerta_visual.short_description = "TIEMPO"

    def cliente_neon(self, obj): return mark_safe(f'<b style="color: #fff; font-size: 0.95rem;">{obj.cliente.upper()}</b>')
    cliente_neon.short_description = "CLIENTE"

    def piso_badge(self, obj):
        color = "#ffcc00" if obj.mesa.piso == "Piso 1" else "#00eaff"
        return mark_safe(f'<span style="background: {color}; color: #000; padding: 2px 6px; border-radius: 3px; font-weight: bold; font-size: 9px;">{obj.mesa.piso.upper()}</span>')
    piso_badge.short_description = "PISO"

    def mesa_visual(self, obj): return mark_safe(f'<span style="color: #fff; border-left: 3px solid #ffcc00; padding-left: 5px;">MESA {obj.mesa.numero}</span>')
    mesa_visual.short_description = "MESA"

    def fecha_hora(self, obj):
        meses = {1:'ENE', 2:'FEB', 3:'MAR', 4:'ABR', 5:'MAY', 6:'JUN', 7:'JUL', 8:'AGO', 9:'SEP', 10:'OCT', 11:'NOV', 12:'DIC'}
        return mark_safe(f'''<div style="display:flex; align-items:center; gap:5px;"><div style="background:#222; border:1px solid #444; border-radius:4px; width:35px; text-align:center;"><div style="background:#ccff00; color:#000; font-size:8px; font-weight:900;">{meses[obj.fecha.month]}</div><div style="color:#fff; font-size:12px; font-weight:bold;">{obj.fecha.day}</div></div><div style="color:#00eaff; font-family:monospace; font-size:1rem; font-weight:bold;">{obj.hora.strftime("%H:%M")}</div></div>''')
    fecha_hora.short_description = "AGENDA"

    def estado_color(self, obj):
        colores = {'Pendiente': '#ffcc00', 'Confirmada': '#28a745', 'Cancelada': '#ff0000'}
        return mark_safe(f'<strong style="color: {colores.get(obj.estado, "#fff")}; font-size: 10px;">● {obj.estado}</strong>')
    estado_color.short_description = "ESTADO"
    
    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        context['mesas_piso1'] = Mesa.objects.filter(piso='Piso 1').order_by('numero')
        context['mesas_piso2'] = Mesa.objects.filter(piso='Piso 2').order_by('numero')
        context['productos_visuales'] = Producto.objects.all().order_by('categoria', 'nombre')
        return super().render_change_form(request, context, add, change, form_url, obj)
